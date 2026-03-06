import pandas as pd
from datetime import datetime
import re
import os
from openpyxl import load_workbook

# ------------------ Case 1 ------------------
def FD_Main(excel_file, output_folder_path):
    # Load the data files
    df_data = pd.read_excel(r"%s" % (excel_file), engine="openpyxl", sheet_name = "Dummy Data")
    df_rate = pd.read_excel(r"%s" % (excel_file), engine="openpyxl", sheet_name = "Rate Card")

    # Step 1: Identify IBG and CBG cases based on CIF_ID
    df_data['CIF_Type'] = df_data['CIF_ID'].astype(str).str[0].map({'1': 'IBG', '2': 'CBG'})

    # Step 2: Calculate age and filter CBG customers aged 60 and above
    today = datetime.today()
    df_data['DATE_OF_BIRTH'] = pd.to_datetime(df_data['DATE_OF_BIRTH'], errors='coerce')
    df_data['Age'] = df_data['DATE_OF_BIRTH'].apply(
        lambda dob: today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day)) if pd.notnull(dob) else None
    )
    cbg_senior_customers = df_data[(df_data['CIF_Type'] == 'CBG') & (df_data['Age'] >= 60)]

    def convert_period_to_days(period_str):
        # Average days per month based on calendar months
        average_days_per_month = {
            1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30,
            7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31
        }

        try:
            parts = period_str.split('/')
            months = int(parts[0].split()[0])
            days = int(parts[1].split()[0])

            # Calculate total days using average month length
            # If months > 12, assume repeating calendar year
            total_days = 0
            for i in range(months):
                month_index = (i % 12) + 1
                total_days += average_days_per_month[month_index]

            return total_days + days
        except Exception as e:
            return None


    cbg_senior_customers['Deposit_Days'] = cbg_senior_customers['DEPOSIT_PERIOD'].apply(convert_period_to_days)

    # Step 4: Match deposit days to rate card range
    def parse_period_to_days(period):
        # Convert time expressions to total days
        def time_to_days(text):
            days = 0
            year_match = re.search(r'(\d+)\s*year', text)
            month_match = re.search(r'(\d+)\s*month', text)
            day_match = re.search(r'(\d+)\s*day', text)

            if year_match:
                days += int(year_match.group(1)) * 365
            if month_match:
                days += int(month_match.group(1)) * 30.44  # average month length
            if day_match:
                days += int(day_match.group(1))
            return int(days)

        # Handle different formats
        if 'to' in period or '& up to' in period:
            parts = re.split(r'to|& up to', period)
            lower = time_to_days(parts[0])
            upper = time_to_days(parts[1])
            return ('range', lower, upper)
        elif 'and above' in period or 'above' in period:
            lower = time_to_days(period)
            return ('above', lower)
        elif 'less than' in period:
            upper = time_to_days(period)
            return ('below', upper)
        else:
            exact = time_to_days(period)
            return ('exact', exact)

    def get_expected_rate(days):
        for _, row in df_rate.iterrows():
            period = row['Time Period']
            try:
                kind, *values = parse_period_to_days(period)
                if kind == 'range':
                    lower, upper = values
                    if lower <= days < upper:
                        return row['Senior Citizen']
                elif kind == 'above':
                    lower = values[0]
                    if days >= lower:
                        return row['Senior Citizen']
                elif kind == 'below':
                    upper = values[0]
                    if days < upper:
                        return row['Senior Citizen']
                elif kind == 'exact':
                    exact = values[0]
                    if days == exact:
                        return row['Senior Citizen']
            except Exception as e:
                continue
        return None

    cbg_senior_customers['Expected_Interest_Rate'] = cbg_senior_customers.apply(
        lambda row: get_expected_rate(row['Deposit_Days']),
        axis=1
    )

    cbg_senior_customers['Expected_Interest_Rate'] = cbg_senior_customers['Expected_Interest_Rate'] * 100

    # Step 5: Identify discrepancies
    cbg_senior_customers['Interest_Rate_Difference'] = (
        cbg_senior_customers['EIT_INTEREST_RATE'].astype(float) -
        cbg_senior_customers['Expected_Interest_Rate'].astype(float)
    ).round(2)

    # Filter discrepancies where the absolute difference is greater than 0.01
    discrepancies = cbg_senior_customers[cbg_senior_customers['Interest_Rate_Difference'].abs() > 0.01]

    # Save results
    file_name = "FD - Focused Sampling Output.xlsx"
    full_path = rf"{output_folder_path}\{file_name}"

    # Ensure the folder exists
    os.makedirs(os.path.dirname(full_path), exist_ok=True)

    # Create a blank DataFrame
    df = pd.DataFrame()

    # Save it to create the file
    df.to_excel(full_path, index=False)

    # Save to a specific sheet in an Excel file
    with pd.ExcelWriter(full_path, engine='openpyxl', mode='a') as writer:
        discrepancies.to_excel(writer, sheet_name='Case_1', index=False)

    # ------------------ Case 2 ------------------

    # Load the data files
    df = pd.read_excel(r"%s" % (excel_file), engine="openpyxl", sheet_name = "Dummy Data")
    rate_card = pd.read_excel(r"%s" % (excel_file), engine="openpyxl", sheet_name = "Rate Card")

    # Step 1: Identify IBG and CBG cases based on CIF_ID
    df['CIF_Type'] = df['CIF_ID'].astype(str).str[0].map({'1': 'IBG', '2': 'CBG'})

    # Step 2: Filter CBG cases with SCHEME_CODE starting with 'ST'
    cbg_df = df[df['CIF_Type'] == 'CBG']
    cbg_df = cbg_df[cbg_df['SCHEME_CODE'].astype(str).str.startswith('ST')]

    # Step 3: Filter cases where age is below 60
    today = datetime.today()
    cbg_df['DATE_OF_BIRTH'] = pd.to_datetime(cbg_df['DATE_OF_BIRTH'], errors='coerce')
    cbg_df['Age'] = cbg_df['DATE_OF_BIRTH'].apply(
        lambda dob: today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day)) if pd.notnull(dob) else None
    )
    cbg_df = cbg_df[cbg_df['Age'] < 60]

    # Step 4: Filter cases with blank STAFF_ID
    cbg_df = cbg_df[cbg_df['sTAFF_ID'].isna()]

    # Step 5: Convert DEPOSIT_PERIOD to total days
    def convert_period_to_days(period_str):
        # Average days per month based on calendar months
        average_days_per_month = {
            1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30,
            7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31
        }

        try:
            parts = period_str.split('/')
            months = int(parts[0].split()[0])
            days = int(parts[1].split()[0])

            # Calculate total days using average month length
            # If months > 12, assume repeating calendar year
            total_days = 0
            for i in range(months):
                month_index = (i % 12) + 1
                total_days += average_days_per_month[month_index]

            return total_days + days
        except Exception as e:
            return None

    cbg_df['Deposit_Days'] = cbg_df['DEPOSIT_PERIOD'].apply(convert_period_to_days)

    # Step 6: Match deposit days to rate card range
    def parse_range(period_str):
        period_str = period_str.lower().strip()
        if 'to less than' in period_str:
            parts = period_str.split('to less than')
            lower = int(re.findall(r'\\d+', parts[0])[0])
            upper = int(re.findall(r'\\d+', parts[1])[0])
            return lower, upper
        elif 'to' in period_str and '&' not in period_str:
            parts = period_str.split('to')
            lower = int(re.findall(r'\\d+', parts[0])[0])
            upper = int(re.findall(r'\\d+', parts[1])[0])
            return lower, upper
        elif 'up to' in period_str:
            upper = int(re.findall(r'\\d+', period_str)[-1])
            return 0, upper
        elif 'and above' in period_str or 'above' in period_str:
            lower = int(re.findall(r'\\d+', period_str)[0])
            return lower, float('inf')
        elif 'less than' in period_str:
            upper = int(re.findall(r'\\d+', period_str)[0])
            return 0, upper
        else:
            days = int(re.findall(r'\\d+', period_str)[0])
            return days, days

    def get_expected_rate(days, is_senior):
        for _, row in rate_card.iterrows():
            lower, upper = parse_range(row['Time Period'])
            if lower <= days <= upper:
                return row['Senior Citizen'] if is_senior else row['Regular']
        return None

    cbg_df['Expected_Interest_Rate'] = cbg_df['Deposit_Days'].apply(
        lambda days: get_expected_rate(days, False)  # False because you've already filtered for age < 60
    )

    cbg_df['Expected_Interest_Rate'] = cbg_df['Expected_Interest_Rate'] * 100

    # Step 7: Identify discrepancies
    cbg_df['Interest_Rate_Difference'] = cbg_df['EIT_INTEREST_RATE'] - cbg_df['Expected_Interest_Rate']
    discrepancies = cbg_df[cbg_df['Interest_Rate_Difference'].abs() > 0.01]

    # Save to a specific sheet in an Excel file
    with pd.ExcelWriter(full_path, engine='openpyxl', mode='a') as writer:
        discrepancies.to_excel(writer, sheet_name='Case_2', index=False)

    # ------------------ Case 4 ------------------

    # Load the Excel file
    df = pd.read_excel(r"%s" % (excel_file), engine="openpyxl", sheet_name = "Dummy Data")

    # Convert the relevant columns to datetime
    df['LCHG_DATE'] = pd.to_datetime(df['LCHG_DATE'], errors='coerce')
    df['VALUE_DATE_OPEN'] = pd.to_datetime(df['VALUE_DATE_OPEN'], errors='coerce')

    # Step 1: Calculate the difference in days
    df['Date_Diff'] = (df['LCHG_DATE'] - df['VALUE_DATE_OPEN']).dt.days

    # Step 2: Select rows where the difference is >= 3
    filtered_df = df[df['Date_Diff'] >= 3]

    # Save to a specific sheet in an Excel file
    with pd.ExcelWriter(full_path, engine='openpyxl', mode='a') as writer:
        filtered_df.to_excel(writer, sheet_name='Case_3', index=False)

    # ------------------ Case 4 ------------------

    # Load the Dummy Data file
    df = pd.read_excel(r"%s" % (excel_file), engine="openpyxl", sheet_name = "Dummy Data")
    full_total = len(df)

    # Step 1: Identify IBG and CBG cases based on CIF_ID
    df['CIF_Type'] = df['CIF_ID'].astype(str).str[0].map({'1': 'IBG', '2': 'CBG'})
    ibg_df = df[df['CIF_Type'] == 'IBG']
    cbg_df = df[df['CIF_Type'] == 'CBG']
    ibg_count = int((len(ibg_df) / full_total) * 1800)
    cbg_count = int((len(cbg_df)/full_total) * 1800)

    # Step 2: Filter rows where VALUE_DATE_INSTRUCTION contains a 10-digit reference number
    df['Ref_10_Digit'] = df['VALUE_DATE_INSTRUCTION'].astype(str).str.extract(r'(\d{10})')
    df = df[df['Ref_10_Digit'].notna()]

    # Step 3: Exclude rows where SCHEME_CODE starts with 'ST'
    df = df[~df['SCHEME_CODE'].astype(str).str.startswith('ST')]

    # Step 4: Exclude rows where age is 60 or above
    df['DATE_OF_BIRTH'] = pd.to_datetime(df['DATE_OF_BIRTH'], errors='coerce')
    today = datetime.today()
    df['Age'] = df['DATE_OF_BIRTH'].apply(
        lambda dob: today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day)) if pd.notnull(dob) else None
    )
    df = df[df['Age'] < 60]

    # Step 5: Proportionally sample 1800 rows between IBG and CBG
    ibg_df = df[df['CIF_Type'] == 'IBG']
    cbg_df = df[df['CIF_Type'] == 'CBG']
    total = len(df)

    sampled_ibg = ibg_df.sample(n=min(ibg_count, len(ibg_df)), random_state=42)
    sampled_cbg = cbg_df.sample(n=min(cbg_count, len(cbg_df)), random_state=42)

    # Combine the sampled data
    sampled_df = pd.concat([sampled_ibg, sampled_cbg])

    # Map the ratio to each row based on CIF_Type
    sampled_df['Proportional_Sample_Size'] = sampled_df['CIF_Type'].map({
        'IBG': ibg_count,
        'CBG': cbg_count
    })

    # Save to a specific sheet in an Excel file
    with pd.ExcelWriter(full_path, engine='openpyxl', mode='a') as writer:
        sampled_df.to_excel(writer, sheet_name='Case_4', index=False)

    # Deleting Sheet 1
    # Load the workbook
    workbook = load_workbook(full_path)

    # Check if 'Sheet1' exists and remove it
    if 'Sheet1' in workbook.sheetnames:
        sheet_to_remove = workbook['Sheet1']
        workbook.remove(sheet_to_remove)

    # Save the workbook
    workbook.save(full_path)

    return full_path